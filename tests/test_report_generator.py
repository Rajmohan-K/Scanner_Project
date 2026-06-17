import pandas as pd
from openpyxl import load_workbook
from reports.report_generator import generate_scan_report


def test_generate_scan_report_handles_market_open_validation():
    stock = {
        "stock": "TEST",
        "score": 42,
        "confidence_pct": 75,
        "final_opportunity_score": 82,
        "opportunity_classification": "High Probability",
        "market_open_analysis": {
            "market_open_price": 100.0,
            "pre_open_change_pct": 1.0,
            "open_to_target_change_pct": 0.5,
        },
        "market_open_validation": {
            "opening_strength_pct": 2.0,
            "order_flow_strength": 10.0,
            "buy_sell_pressure": 25.0,
            "final_trade_quality_score": 88.0,
            "opportunity_classification": "Good Opportunity",
            "premarket_confidence_score": 70.0,
            "market_open_confirmation_score": 65.0,
            "price_acceptance_above_key_levels": 80.0,
            "price_rejection_below_key_levels": 20.0,
            "relative_volume_increase": 120.0,
            "volume_change_from_premarket_volume": 15.0,
        },
    }

    report_path = generate_scan_report([stock])
    assert report_path is not None
    assert report_path.endswith(".xlsx")


def test_generate_scan_report_writes_tiered_sheets():
    all_rows = [
        {"stock": "AAA.NS", "live_price": 100, "last_close": 99, "coarse_quality": 70},
        {"stock": "BBB.NS", "live_price": 90, "last_close": 91, "coarse_quality": 60},
    ]
    filtered_rows = [{"stock": "AAA.NS", "score": 40, "ml_probability": 72}]
    top_rows = [{"stock": "AAA.NS", "score": 40, "ml_probability": 72, "premarket_action": "BUY"}]

    report_path = generate_scan_report(
        top_rows,
        all_results=all_rows,
        filtered_results=filtered_rows,
        top_results=top_rows,
    )

    workbook = load_workbook(report_path, read_only=True)
    assert "All_Stocks_Live_Data" in workbook.sheetnames
    assert "Filtered_150" in workbook.sheetnames
    assert "Top_25" in workbook.sheetnames
    assert workbook["All_Stocks_Live_Data"].max_row == 3
